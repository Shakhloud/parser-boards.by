"""
Парсер рекламных поверхностей с сайта boards.by

Алгоритм:
1. Загружаем главную страницу https://boards.by/
2. Извлекаем JSON из атрибута data-banners элемента #map
3. Для каждой конструкции:
   - Нормализуем display_type и construction_format
   - Разбиваем стороны (PROPERTY_SIDE_VALUE) на отдельные записи
4. Сохраняем в БД (PostgreSQL)
5. Экспортируем в result.json и result.xlsx

Запуск:
    python parse.py
"""

import json
import re
import logging
import time
from os import getenv
from typing import Optional
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from sqlalchemy.orm import sessionmaker

from app.database import engine, Base
from app.models import Construction, ConstructionSide
from app.schemas import (
    normalize_display_type,
    normalize_construction_format,
    normalize_size,
    ExportResult,
    ConstructionSideExport,
)

# Загружаем переменные из .env
load_dotenv()

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# Конфигурация
BASE_URL = "https://boards.by"
TARGET_SIDES = int(getenv("TARGET_SIDES", 100))
MAX_CONSTRUCTIONS = int(getenv("MAX_CONSTRUCTIONS", 100))
REQUEST_DELAY = float(getenv("REQUEST_DELAY", 0.02))
OUTPUT_JSON = getenv("OUTPUT_JSON", "result.json")
OUTPUT_XLSX = getenv("OUTPUT_XLSX", "result.xlsx")


def fetch_main_page() -> str:
    """Загружает главную страницу boards.by."""
    logger.info(f"Загружаю {BASE_URL} ...")
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    }
    response = requests.get(BASE_URL, headers=headers, timeout=30)
    response.raise_for_status()
    logger.info(f"Страница загружена: {len(response.text)} байт")
    return response.text


def extract_banners_data(html: str) -> dict:
    """Извлекает JSON с данными о конструкциях из HTML."""
    soup = BeautifulSoup(html, "html.parser")
    map_element = soup.find("div", id="map")
    
    if not map_element:
        raise ValueError("Не найден элемент #map на странице")
    
    data_banners = map_element.get("data-banners")
    if not data_banners:
        raise ValueError("Атрибут data-banners пуст или отсутствует")
    
    banners = json.loads(data_banners)
    logger.info(f"Извлечено {len(banners)} конструкций из data-banners")
    return banners


def parse_sides(side_value: Optional[str]) -> list[str]:
    """
    Разбивает строку сторон на список.
    
    Примеры:
        "А"         → ["А"]
        "А,В"       → ["А", "В"]
        "А1,А2,А3"  → ["А1", "А2", "А3"]
        None        → []
    """
    if not side_value:
        return []
    
    # Разделяем по запятой, убираем пробелы
    sides = [s.strip() for s in side_value.split(",")]
    return [s for s in sides if s]


def fetch_card_page(code: str) -> Optional[str]:
    """Загружает страницу карточки по code (URL-slug)."""
    url = f"{BASE_URL}/banner/{code}/"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    }
    try:
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        return response.text
    except Exception as e:
        logger.warning(f"Ошибка загрузки карточки {url}: {e}")
        return None


def parse_card_details(html: str) -> dict:
    """
    Парсит детальную информацию из карточки объекта.
    
    Возвращает:
        {
            "size": "6x3" или None,
            "display_type": "Призматрон" или None,
        }
    """
    if not html:
        return {"size": None, "display_type": None}
    
    soup = BeautifulSoup(html, "html.parser")
    result = {"size": None, "display_type": None}
    
    # Ищем все .dl_element блоки
    dl_elements = soup.find_all("div", class_="dl_element")
    
    for dl in dl_elements:
        spans = dl.find_all("span")
        if len(spans) < 2:
            continue
        
        label = spans[0].get_text(strip=True)
        value = spans[1].get_text(strip=True)
        
        if label == "Формат":
            # Нормализуем размер: "6*3м" → "6x3"
            result["size"] = normalize_size(value)
        elif label == "Тип":
            # Нормализуем тип отображения
            result["display_type"] = normalize_display_type(value)
    
    
    return result


def parse_single_banner(banner_data: dict, fetch_details: bool = False) -> tuple[dict, list[dict]]:
    """
    Парсит одну запись из data-banners.
    
    Параметры:
        banner_data: сырые данные из data-banners
        fetch_details: если True, загружает карточку для получения size и display_type
    
    Возвращает:
        (construction_data, list_of_side_data)
    """
    # Извлекаем базовые данные
    gid = banner_data.get("ID", "")
    name = banner_data.get("NAME", "")
    code = banner_data.get("CODE", "")
    
    # Координаты (обратите внимание: LONGITUDE = долгота, LATITUDE = широта)
    lon_str = banner_data.get("PROPERTY_LONGITUDE_VALUE", "0")
    lat_str = banner_data.get("PROPERTY_LATITUDE_VALUE", "0")
    
    try:
        lon = float(lon_str) if lon_str else 0.0
        lat = float(lat_str) if lat_str else 0.0
    except ValueError:
        lon, lat = 0.0, 0.0
    
    # Тип конструкции (из PROPERTY_TYPE_VALUE) — базовое значение
    raw_type = banner_data.get("PROPERTY_TYPE_VALUE", "")
    display_type = normalize_display_type(raw_type)
    construction_format = normalize_construction_format(raw_type)
    
    # Адрес из NAME или PROPERTY_REGION_VALUE
    region = banner_data.get("PROPERTY_REGION_VALUE", "")
    address = name if name else region
    
    # Стороны
    raw_sides = banner_data.get("PROPERTY_SIDE_VALUE", "")
    sides = parse_sides(raw_sides)
    
    # URL карточки
    source_url = f"{BASE_URL}/banner/{code}/"
    
    # Загружаем детальную информацию из карточки
    card_size = None
    card_display_type = None
    
    if fetch_details and code:
        time.sleep(REQUEST_DELAY)  # Не DDOS'им сайт
        card_html = fetch_card_page(code)
        if card_html:
            card_details = parse_card_details(card_html)
            card_size = card_details["size"]
            card_display_type = card_details["display_type"]
            # Если на карточке точнее указан display_type — используем его
            if card_display_type:
                display_type = card_display_type
    
    # Данные конструкции
    construction = {
        "gid": str(gid),
        "address": address,
        "lon": lon,
        "lat": lat,
        "construction_format": construction_format,
        "display_type": display_type,
        "lighting": None,  # На сайте нет данных о подсветке
        "source_url": source_url,
        "raw_data": json.dumps(banner_data, ensure_ascii=False),
    }
    
    # Если сторон нет — создаём одну запись с name = gid
    if not sides:
        side = {
            "name": str(gid),
            "size": card_size,
            "material": None,  # На сайте нет данных о материале
            "display_type": display_type,
            "lighting": None,
            "source_url": source_url,
            "raw_data": None,
        }
        return construction, [side]
    
    # Данные сторон
    sides_data = []
    for side_name in sides:
        side = {
            "name": side_name,
            "size": card_size,  # Размер одинаковый для всех сторон конструкции
            "material": None,
            "display_type": display_type,
            "lighting": None,
            "source_url": f"{source_url}?side={side_name}",
            "raw_data": None,
        }
        sides_data.append(side)
    
    return construction, sides_data


def parse_boards_by(min_sides: int = TARGET_SIDES, fetch_details: bool = True, max_constructions: int = MAX_CONSTRUCTIONS) -> tuple[list[dict], list[tuple[str, dict]]]:
    """
    Главная функция парсинга.
    
    Параметры:
        min_sides: минимальное количество сторон для сбора
        fetch_details: если True, загружает карточки для size и display_type
        max_constructions: максимум конструкций для парсинга (0 = все)
    
    Возвращает:
        (constructions, sides_with_gid)
        где sides_with_gid = [(gid, side_data), ...]
    """
    html = fetch_main_page()
    banners = extract_banners_data(html)
    
    constructions = []
    all_sides = []
    total = len(banners)
    
    # Ограничиваем количество конструкций
    if max_constructions > 0:
        banners_to_parse = dict(list(banners.items())[:max_constructions])
        logger.info(f"Ограничение: парсим только {len(banners_to_parse)} из {total} конструкций")
    else:
        banners_to_parse = banners
    
    for idx, (banner_id, banner_data) in enumerate(banners_to_parse.items(), 1):
        try:
            # Показываем прогресс
            if idx % 25 == 0 or idx == len(banners_to_parse):
                logger.info(f"Прогресс: {idx}/{len(banners_to_parse)} конструкций, {len(all_sides)} сторон")
            
            construction, sides = parse_single_banner(banner_data, fetch_details=fetch_details)
            constructions.append(construction)
            
            for side in sides:
                all_sides.append((construction["gid"], side))
                
        except Exception as e:
            logger.warning(f"Ошибка парсинга баннера {banner_id}: {e}")
            continue
    
    logger.info(f"Распаршено: {len(constructions)} конструкций, {len(all_sides)} сторон")
    
    if len(all_sides) < min_sides:
        logger.warning(
            f"Собрано только {len(all_sides)} сторон, а нужно минимум {min_sides}"
        )
    
    return constructions, all_sides


def save_to_database(
    constructions: list[dict],
    sides_with_gid: list[tuple[str, dict]],
) -> int:
    """
    Сохраняет данные в БД.
    
    Возвращает количество сохранённых сторон.
    """
    Session = sessionmaker(bind=engine)
    session = Session()
    
    try:
        # Создаём мапу gid → construction_id
        gid_to_id = {}
        
        for cons_data in constructions:
            # Проверяем, есть ли уже такая конструкция
            existing = session.query(Construction).filter(
                Construction.gid == cons_data["gid"]
            ).first()
            
            if existing:
                gid_to_id[cons_data["gid"]] = existing.id
                continue
            
            construction = Construction(**cons_data)
            session.add(construction)
            session.flush()  # Получаем ID
            gid_to_id[cons_data["gid"]] = construction.id
        
        # Сохраняем стороны
        saved_sides = 0
        for gid, side_data in sides_with_gid:
            construction_id = gid_to_id.get(gid)
            if not construction_id:
                logger.warning(f"Конструкция с gid={gid} не найдена")
                continue
            
            # Проверяем дубликаты
            existing = session.query(ConstructionSide).filter(
                ConstructionSide.construction_id == construction_id,
                ConstructionSide.name == side_data["name"],
            ).first()
            
            if existing:
                continue
            
            side = ConstructionSide(
                construction_id=construction_id,
                **side_data,
            )
            session.add(side)
            saved_sides += 1
        
        session.commit()
        logger.info(f"Сохранено в БД: {len(gid_to_id)} конструкций, {saved_sides} сторон")
        return saved_sides
        
    except Exception as e:
        session.rollback()
        logger.error(f"Ошибка сохранения в БД: {e}")
        raise
    finally:
        session.close()


def load_from_database() -> list[ConstructionSideExport]:
    """
    Загружает данные из БД в формат выгрузки.
    
    Возвращает список ConstructionSideExport.
    """
    Session = sessionmaker(bind=engine)
    session = Session()
    
    try:
        # Загружаем все стороны с присоединением конструкций
        sides = (
            session.query(ConstructionSide)
            .join(Construction)
            .all()
        )
        
        export_data = []
        for side in sides:
            cons = side.construction
            export_item = ConstructionSideExport(
                gid=cons.gid,
                address=cons.address,
                name=side.name,
                lon=cons.lon,
                lat=cons.lat,
                construction_format=cons.construction_format,
                display_type=side.display_type or cons.display_type,
                lighting=side.lighting if side.lighting is not None else cons.lighting,
                size=normalize_size(side.size),
                material=side.material,
            )
            export_data.append(export_item)
        
        logger.info(f"Загружено из БД: {len(export_data)} сторон")
        return export_data
        
    finally:
        session.close()


def export_to_json(data: list[ConstructionSideExport], output_path: str = "result.json") -> None:
    """Экспортирует данные в JSON файл."""
    result = ExportResult(construction_sides=data)
    
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result.model_dump(), f, ensure_ascii=False, indent=2)
    
    logger.info(f"Экспорт в JSON: {output_path} ({len(data)} записей)")


def export_to_xlsx(data: list[ConstructionSideExport], output_path: str = "result.xlsx") -> None:
    """Экспортирует данные в Excel файл."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "construction_sides"
    
    # Заголовки
    headers = [
        "gid", "address", "name", "lon", "lat",
        "construction_format", "display_type",
        "lighting", "size", "material"
    ]
    ws.append(headers)
    
    # Данные
    for item in data:
        ws.append([
            item.gid,
            item.address,
            item.name,
            item.lon,
            item.lat,
            item.construction_format,
            item.display_type,
            "Да" if item.lighting is True else "Нет" if item.lighting is False else "",
            item.size,
            item.material,
        ])
    
    # Автоширина колонок
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        max_length = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_num, max_col=col_num):
            cell_value = str(row[0].value or "")
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    wb.save(output_path)
    logger.info(f"Экспорт в XLSX: {output_path} ({len(data)} записей)")


def main():
    """Главная функция."""
    logger.info("=" * 60)
    logger.info("Парсер boards.by — запуск")
    logger.info("=" * 60)
    
    # 1. Парсим данные
    logger.info("Шаг 1: Парсинг данных с boards.by")
    constructions, sides = parse_boards_by(min_sides=TARGET_SIDES)
    
    # 2. Сохраняем в БД
    logger.info("Шаг 2: Сохранение в БД")
    save_to_database(constructions, sides)
    
    # 3. Загружаем из БД
    logger.info("Шаг 3: Загрузка из БД для экспорта")
    export_data = load_from_database()
    
    # 4. Экспортируем
    logger.info("Шаг 4: Экспорт в файлы")
    export_to_json(export_data, OUTPUT_JSON)
    export_to_xlsx(export_data, OUTPUT_XLSX)
    
    logger.info("=" * 60)
    logger.info(f"Готово! Собрано {len(export_data)} сторон")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
